import sqlite3
from openpyxl import load_workbook

conn = sqlite3.connect('history.db')
wb = load_workbook('study_index_2.xlsx')
ws = wb['Sheet1']

conn.execute("DROP TABLE topics");
conn.execute("create table if not exists topics (id int NOT NULL PRIMARY KEY,subject text, topic text, link text, page int)")

for i in range(1,44):
    temp_str = "insert into topics (id,subject, topic, link, page) values ('{0}', '{1}', '{2}', '{3}','{4}')".format(ws.cell(i,1).value, ws.cell(i,2).value, ws.cell(i,3).value, ws.cell(i,4).value, ws.cell(i,5).value)
    conn.execute(temp_str)

conn.commit()

content = conn.execute("select * from topics")
for i in content:
    print(i)
